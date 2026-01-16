from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime


def generate_customers_excel(customers):
    """
    Gera um arquivo Excel com dados dos clientes
    
    Args:
        customers: Lista de dicionários com dados dos clientes
    """
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes"
    
    # Estilos
    header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = ["ID", "Nome", "Email", "Telefone", "Data de Cadastro"]
    ws.append(headers)
    
    # Aplicar estilos ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Adicionar dados
    for customer in customers:
        ws.append([
            customer.get('id', ''),
            customer.get('name', ''),
            customer.get('email', ''),
            customer.get('phone', ''),
            customer.get('created_at', ''),
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20
    
    # Aplicar bordas aos dados
    for row in ws.iter_rows(min_row=2, max_row=len(customers) + 1, min_col=1, max_col=5):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Salvar em memória
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def generate_budgets_excel(budgets, customers):
    """
    Gera um arquivo Excel com dados dos orçamentos
    
    Args:
        budgets: Lista de dicionários com dados dos orçamentos
        customers: Lista de dicionários com dados dos clientes
    """
    
    # Criar workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamentos"
    
    # Estilos
    header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = ["ID", "Título", "Cliente", "Subtotal", "Desconto (%)", "Total Final", "Status", "Data"]
    ws.append(headers)
    
    # Aplicar estilos ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Adicionar dados
    for budget in budgets:
        # Encontrar nome do cliente
        customer_name = "-"
        if budget.get('customer_id'):
            customer = next((c for c in customers if c.get('id') == budget.get('customer_id')), None)
            if customer:
                customer_name = customer.get('name', '-')
        
        ws.append([
            budget.get('id', ''),
            budget.get('title', ''),
            customer_name,
            f"R$ {budget.get('subtotal_amount', 0):.2f}",
            f"{budget.get('discount_percent', 0)}%",
            f"R$ {budget.get('final_amount', 0):.2f}",
            budget.get('status', '').upper(),
            budget.get('created_at', ''),
        ])
    
    # Ajustar largura das colunas
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    
    # Aplicar bordas aos dados
    for row in ws.iter_rows(min_row=2, max_row=len(budgets) + 1, min_col=1, max_col=8):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="left", vertical="center")
    
    # Salvar em memória
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer


def generate_monthly_report_excel(budgets, customers):
    """
    Gera um relatório mensal em Excel com análises
    
    Args:
        budgets: Lista de dicionários com dados dos orçamentos
        customers: Lista de dicionários com dados dos clientes
    """
    
    # Criar workbook com múltiplas abas
    wb = Workbook()
    
    # ===== ABA 1: Orçamentos =====
    ws1 = wb.active
    ws1.title = "Orçamentos"
    
    # Estilos
    header_fill = PatternFill(start_color="1f2937", end_color="1f2937", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos
    headers = ["ID", "Título", "Cliente", "Subtotal", "Desconto (%)", "Total Final", "Status", "Data"]
    ws1.append(headers)
    
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Adicionar dados
    for budget in budgets:
        customer_name = "-"
        if budget.get('customer_id'):
            customer = next((c for c in customers if c.get('id') == budget.get('customer_id')), None)
            if customer:
                customer_name = customer.get('name', '-')
        
        ws1.append([
            budget.get('id', ''),
            budget.get('title', ''),
            customer_name,
            f"R$ {budget.get('subtotal_amount', 0):.2f}",
            f"{budget.get('discount_percent', 0)}%",
            f"R$ {budget.get('final_amount', 0):.2f}",
            budget.get('status', '').upper(),
            budget.get('created_at', ''),
        ])
    
    # Ajustar largura
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws1.column_dimensions[col].width = 20
    
    # ===== ABA 2: Análise Resumida =====
    ws2 = wb.create_sheet("Análise")
    
    # Cálculos
    total_budgets = len(budgets)
    approved = len([b for b in budgets if b.get('status') == 'approved'])
    draft = len([b for b in budgets if b.get('status') == 'draft'])
    rejected = len([b for b in budgets if b.get('status') == 'rejected'])
    total_revenue = sum(b.get('final_amount', 0) for b in budgets)
    
    # Dados
    ws2.append(["Métrica", "Valor"])
    ws2.append(["Total de Orçamentos", total_budgets])
    ws2.append(["Orçamentos Aprovados", approved])
    ws2.append(["Orçamentos em Rascunho", draft])
    ws2.append(["Orçamentos Rejeitados", rejected])
    ws2.append(["Faturamento Total", f"R$ {total_revenue:.2f}"])
    ws2.append(["Ticket Médio", f"R$ {total_revenue / total_budgets:.2f}" if total_budgets > 0 else "R$ 0.00"])
    
    # Estilos
    for row in ws2.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
        for cell in row:
            cell.border = border
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 20
    
    # Salvar em memória
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer
